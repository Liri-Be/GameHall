import math
import random
import sys
from time import time, sleep
import pygame
from openpyxl import load_workbook
import pandas

pygame.font.init()


class Game:
    def __init__(self, name, screen):
        """
        init objects
        :param name: the name of the game
        :type name: str
        :param screen: the pygame screen of the game
        :type screen: pygame.Surface
        """
        self.name = name
        self.rounds = 0  # number of rounds for a win
        self.points = 0
        self.screen = screen  # pygame screen
        self.screen_size = 700
        self.bg_color = (173, 216, 230)
        self.font = pygame.font.SysFont("arial bold", 50)  # font for pygame
        self.stat = {'wins': 0, 'ties': 0, 'turns': [], 'time': [], 'choosing': []}
        self.lead_board_path = r'files\{0}\lead_board.xlsx'.format(name)

    def playGame(self):
        """
        handles the whole game - play each round until winning the desired amount, and present statistics
        :return: None
        """
        curr_round = 0
        # while self.stat['wins'] != self.rounds:
        for _ in range(5):  # play 5 rounds
            self.drawBoard()
            start = time()
            won, turns, tie = self.playOneRound()
            finish = time()
            self.resetAll()
            curr_round += 1
            self.statistics_round(won, turns, finish - start, 0, curr_round, tie)

        self.statistics_game(curr_round)
        return

    def playOneRound(self):
        return True, 0, False

    def checkWin(self):
        pass

    def checkTie(self):
        pass

    def drawBoard(self):
        pass

    def resetAll(self):
        pass

    def chooseAmountRounds(self):
        """
        handles the choosing amount of rounds to declare a win
        :return: None
        """
        # present choose amount of rounds
        # present background
        rnd_screen = pygame.image.load(r'photos\General\rounds.png')
        self.screen.blit(rnd_screen, (0, 0))
        pygame.display.update()

        # present input text rectangle
        input_rect = pygame.Rect(200, 375, 300, 65)
        pygame.draw.rect(self.screen, (34, 177, 36), input_rect)

        # variables for input
        done_to_enter = False
        user_input = ""
        digit = True
        empty = False
        zero = False

        # get the input
        while not done_to_enter:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:  # if press exit - leave
                    pygame.quit()
                    sys.exit()

                elif event.type == pygame.KEYDOWN:  # if typing record the input if valid
                    error_rect = pygame.Rect(0, 470, 700, 300)
                    pygame.draw.rect(self.screen, self.bg_color, error_rect)
                    pygame.display.flip()

                    if event.key == pygame.K_BACKSPACE:  # backspace - delete the last one
                        user_input = user_input[:-1]
                        pygame.draw.rect(self.screen, (34, 177, 36), input_rect)
                        user_txt = self.font.render(user_input, False, (0, 0, 0), (34, 177, 36))
                        self.screen.blit(user_txt, (200, 375))
                        pygame.display.flip()

                    elif event.key == pygame.K_RETURN:  # pressed enter
                        if user_input == "":
                            empty = True
                        else:
                            done_to_enter = True

                    else:  # pressed other key
                        empty = False
                        if event.unicode.isdigit():  # if it is a digit
                            digit = True
                            if int(event.unicode) == 0 and user_input == "":  # input is zero
                                zero = True
                            else:
                                user_input += event.unicode

                                zero = False
                        else:  # not a digit
                            digit = False

            if empty:  # try to enter empty amount
                error_msg = self.font.render("Not valid, enter at least one number.",
                                             False, (255, 0, 0), self.bg_color)
                self.screen.blit(error_msg, (52, 475))

            elif zero and digit:  # if it is zero
                error_msg = self.font.render("Not valid, number of rounds", False, (255, 0, 0), self.bg_color)
                self.screen.blit(error_msg, (127, 475))
                error_msg = self.font.render("should be bigger than 0.", False, (255, 0, 0), self.bg_color)
                self.screen.blit(error_msg, (150, 515))

            elif not digit:  # if we got not digit
                error_msg = self.font.render("Not valid, only numbers.", False, (255, 0, 0), self.bg_color)
                self.screen.blit(error_msg, (150, 475))

            else:  # delete the error msg if we have valid input
                error_rect = pygame.Rect(0, 470, 700, 300)
                pygame.draw.rect(self.screen, self.bg_color, error_rect)

            # present the input on the screen
            user_txt = self.font.render(user_input, False, (0, 0, 0), (34, 177, 36))
            self.screen.blit(user_txt, (200, 375))
            pygame.display.flip()

        self.rounds = int(user_input)
        return

    def statistics_round(self, won, turns, round_time, choosing, curr_round, tie):
        """
        calculate and present the statistics of each round of the game
        :param won: whether the user won the round
        :type won: bool
        :param turns: number of turns this round took
        :type turns: int
        :param round_time: the time the round took
        :type round_time: float
        :param choosing: the average time it took the user to choose an action
        :type choosing: float
        :param curr_round: number of the current round
        :type curr_round: int
        :param tie: whether the round ended with a tie
        :type tie: bool
        :return: the msg of the statistics, string
        """
        self.stat['time'].append(round_time)  # save time
        self.stat['turns'].append(turns)  # save turns
        self.stat['choosing'].append(choosing)
        if won:
            self.stat['wins'] += 1  # remember wins of user
        if tie:
            self.stat['ties'] += 1  # remember ties

        avg_wins = self.stat['wins'] / curr_round  # calc avg wins per round
        avg_turns = sum(self.stat['turns']) / curr_round  # calc avg turns per round

        # make msg after one round
        msg = "Here are the round statistics:\n"
        msg += "You won: " + str(self.stat['wins']) + " times\n"
        msg += "The Computer won: " + str(curr_round - self.stat['ties'] - self.stat['wins']) + " times\n"
        msg += "Amount of turns for this round: " + str(turns) + "\n"
        msg += "This round took you " + str(format(self.stat['time'][curr_round - 1], '.3f')) + " seconds\n"
        # msg += "You chose where to place the token in " + str(format(self.stat['choosing'][curr_round - 1], '.3f'))
        # msg += " seconds in average\n"
        msg += "Average amount of turns per round: " + str(format(avg_turns, '.3f')) + "\n"
        msg += "Average amount of wins per round: " + str(format(avg_wins, '.3f')) + "\n\n"

        # present statistics
        self.screen.fill(self.bg_color)
        pygame.display.update()
        sleep(0.5)
        stat_msg = pygame.font.SysFont("arial bold", 60).render("Round ended!", False, (0, 0, 0), self.bg_color)
        self.screen.blit(stat_msg, (210, 20))

        # win msg
        if won and not tie:
            msg_win = "You won! Congratulations!"
            msg_win_color = (50, 190, 60)
            prefix = 0
        elif tie:
            msg_win = "It is a tie! Try again."
            msg_win_color = (225, 211, 12)
            prefix = 83
        else:
            msg_win = "Computer won! Try again."
            msg_win_color = (255, 37, 37)
            prefix = 15
        stat_msg = pygame.font.SysFont("arial bold", 60).render(msg_win, False, msg_win_color, self.bg_color)
        self.screen.blit(stat_msg, (70 + prefix, 65))

        # points msg
        point_msg = pygame.font.SysFont("arial bold", 50).render("You've got {0} points!".format(self.points),
                                                                 False, (0, 0, 0), self.bg_color)
        self.screen.blit(point_msg, (185, 115))

        # the stats
        for i in range(len(msg.split('\n')) - 1):
            stat_msg = pygame.font.SysFont("arial", 38).render(msg.split('\n')[i], False, (0, 0, 0), self.bg_color)
            k = 10 if i == 0 else 50
            self.screen.blit(stat_msg, (k, i * 50 + 172))

        self.screen.blit(self.font.render("Press the mouse to continue", False, (0, 0, 0), self.bg_color), (112, 550))
        pygame.display.flip()  # display the msg

        while True:  # present until pressed mouse
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEBUTTONDOWN:
                    break
            else:
                continue
            break
        print(msg)
        return msg

    def statistics_game(self, rounds):
        """
        calculate and present the statistics of the whole game
        :param rounds: number of rounds the game took
        :type rounds: int
        :return: the msg of the statistics, string
        """
        avg_wins = self.stat['wins'] / rounds  # calc avg wins per round
        avg_turns = sum(self.stat['turns']) / rounds  # calc avg turns per round
        avg_time = sum(self.stat['time']) / rounds  # calc avg time per round
        msg = "Here are the game statistics:\n"
        msg += "You won: " + str(self.stat['wins']) + " times\n"
        msg += "The Computer won: " + str(rounds - self.stat['ties'] - self.stat['wins']) + " times\n"
        msg += "Average amount of time per round: " + str(format(avg_time, '.3f')) + "s\n"
        # msg += "You choose where to place the token in " + str(format(self.stat['choosing'][rounds - 1], '.3f'))
        # msg += " seconds in average\n"
        msg += "Average amount of turns per round: " + str(format(avg_turns, '.3f')) + "\n"
        msg += "Average amount of wins per round: " + str(format(avg_wins, '.3f')) + "\n\n"

        # present statistics
        self.screen.fill(self.bg_color)
        pygame.display.update()
        sleep(0.5)

        # the stats
        stat_msg = pygame.font.SysFont("arial bold", 60).render("Game ended!", False, (0, 0, 0), self.bg_color)
        self.screen.blit(stat_msg, (210, 20))

        # points msg
        point_msg = pygame.font.SysFont("arial bold", 50).render("You've got {0} points!".format(self.points),
                                                                 False, (0, 0, 0), self.bg_color)
        self.screen.blit(point_msg, (185, 70))

        # stat msg
        for i in range(len(msg.split('\n')) - 1):
            stat_msg = pygame.font.SysFont("arial", 38).render(msg.split('\n')[i], False, (0, 0, 0), self.bg_color)
            k = 10 if i == 0 else 50
            self.screen.blit(stat_msg, (k, i * 50 + 127))

        self.screen.blit(self.font.render("Press the mouse to continue", False, (0, 0, 0), self.bg_color), (112, 550))
        pygame.display.flip()  # display the msg

        while True:  # present until pressed mouse
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEBUTTONDOWN:
                    break
            else:
                continue
            break
        print(msg)
        return msg

    def saveRecord(self):
        """
        taking the nickname of the user from the user and checking if it's a valid name
        :return: the nickname of the user, str
        """
        # ask if the user want to be in the lead board
        # present background
        rnd_screen = pygame.image.load(r'photos\General\save_record.png')
        self.screen.blit(rnd_screen, (0, 0))
        pygame.display.update()

        # wait for input from user
        choose = False
        want_lead = False  # want to be in the lead board or not

        while not choose:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:  # leave
                    pygame.quit()
                    sys.exit()
                elif event.type == pygame.MOUSEBUTTONDOWN:  # user pressed mouse
                    x_pos, y_pos = pygame.mouse.get_pos()
                    if 100 < x_pos < 500:  # in the x-range
                        if 327 < y_pos < 457:  # yes
                            print("Yes")
                            want_lead = True
                            choose = True

                        elif 457 < y_pos < 617:  # no
                            want_lead = False
                            choose = True

        if want_lead:  # want to save
            # start to save the name
            # present background
            rnd_screen = pygame.image.load(r'photos\General\ask_name.png')
            self.screen.blit(rnd_screen, (0, 0))
            pygame.display.update()

            # present input text rectangle
            input_rect = pygame.Rect(200, 375, 300, 65)
            pygame.draw.rect(self.screen, (34, 177, 36), input_rect)

            # variables for input
            done_to_enter = False
            user_input = ""
            digit = True
            empty = False

            # get the input
            while not done_to_enter:
                for event in pygame.event.get():
                    if event.type == pygame.QUIT:  # if press exit - leave
                        pygame.quit()
                        sys.exit()

                    elif event.type == pygame.KEYDOWN:  # if typing record the input if valid
                        error_rect = pygame.Rect(0, 470, 700, 300)
                        pygame.draw.rect(self.screen, self.bg_color, error_rect)
                        pygame.display.flip()

                        if event.key == pygame.K_BACKSPACE:  # backspace - delete the last one
                            user_input = user_input[:-1]
                            pygame.draw.rect(self.screen, (34, 177, 36), input_rect)
                            user_txt = self.font.render(user_input, False, (0, 0, 0), (34, 177, 36))
                            self.screen.blit(user_txt, (200, 375))
                            pygame.display.flip()

                        elif event.key == pygame.K_RETURN:  # pressed enter
                            if user_input == "":
                                empty = True
                            else:
                                done_to_enter = True

                        else:  # pressed other key
                            empty = False
                            if event.unicode.isalpha():  # if it is a letter
                                digit = True
                                user_input += event.unicode
                            else:  # not a letter
                                digit = False

                if empty:  # try to enter empty amount
                    error_msg = self.font.render("Not valid, enter at least one letter.",
                                                 False, (255, 0, 0), self.bg_color)
                    self.screen.blit(error_msg, (52, 475))

                elif not digit:  # if we got not digit
                    error_msg = self.font.render("Not valid, only letters.", False, (255, 0, 0), self.bg_color)
                    self.screen.blit(error_msg, (150, 475))

                else:  # delete the error msg if we have valid input
                    error_rect = pygame.Rect(0, 470, 700, 300)
                    pygame.draw.rect(self.screen, self.bg_color, error_rect)

                # present the input on the screen
                user_txt = self.font.render(user_input, False, (0, 0, 0), (34, 177, 36))
                self.screen.blit(user_txt, (200, 375))
                pygame.display.flip()

            return user_input
        return None  # does not want to save

    def presentRecordTable(self):
        """
        present the lead board on screen - top 5,
        save the name of the user if they want to be saved and the user place in the lead board
        :return: None
        """
        # save the name and sort
        name = self.saveRecord()
        if name is not None:  # if the user want to be saved - save its data
            # open the Excel
            workbook = load_workbook(filename=self.lead_board_path)
            sheet = workbook['Sheet']
            # check if already in the lead board
            exist = False
            row = 1
            names = sheet['A']
            for i in range(1, len(names)):
                if names[i].value == name:
                    exist = True
                    row = i + 1
                    break

            # write new data to our excel
            if exist:
                sheet["B{0}".format(row)].value = self.points
                sheet["C{0}".format(row)].value = sum(self.stat['time'])
            else:
                new_data = [name, self.points, sum(self.stat['time'])]
                print(new_data)
                sheet.append(new_data)
            workbook.save(filename=self.lead_board_path)

        # sort by points and time
        workbook = pandas.read_excel(self.lead_board_path)
        workbook = workbook.sort_values(["Points", "Time(s)"], ascending=[False, True])
        workbook.to_excel(r'files\{0}\lead_sorted.xlsx'.format(self.name), sheet_name="Sheet")

        # present the table
        self.screen.fill(self.bg_color)
        pygame.draw.line(self.screen, (0, 0, 0), (265, 145), (265, 595), 7)
        pygame.draw.line(self.screen, (0, 0, 0), (435, 145), (435, 595), 7)
        for i in range(5):
            pygame.draw.line(self.screen, (0, 0, 0), (95, 220 + i*75), (605, 220 + i*75), 7)

        # present the data
        workbook = load_workbook(filename=r'files\{0}\lead_sorted.xlsx'.format(self.name))
        sheet = workbook['Sheet']
        # find top 5
        data = []
        for i in range(1, min(sheet.max_row, 6) + 1):
            row = []
            for j in range(2, sheet.max_column + 1):
                row.append(sheet.cell(row=i, column=j).value)
            data.append(tuple(row))
        print(data)
        # put on screen
        prefix_y = 0
        for user_data in data:
            user_name, user_points, user_time = user_data
            user_points = str(user_points)
            prefix_x_time = 0
            try:  # if it is the time present only 3 digits after the point
                user_time = str(format(user_time, '.3f'))
                prefix_x_time = 20  # to place in the center
            except ValueError:  # if it is the headline
                pass
            prefix_x_points = 45 if user_points.isdigit() else 0  # to place in the center
            self.screen.blit(self.font.render(user_name, False, (0, 0, 0), self.bg_color), (120, 170 + 75*prefix_y))
            self.screen.blit(self.font.render(user_points, False, (0, 0, 0), self.bg_color),
                             (295 + prefix_x_points, 170 + 75*prefix_y))
            self.screen.blit(self.font.render(user_time, False, (0, 0, 0), self.bg_color),
                             (460 + prefix_x_time, 170 + 75*prefix_y))
            prefix_y += 1  # to place in the current line

        # headlines
        self.screen.blit(pygame.font.SysFont("ariel bold", 80).render("Lead board", False,
                                                                      (0, 0, 0), self.bg_color), (195, 30))
        self.screen.blit(self.font.render("Press the mouse to continue", False, (0, 0, 0), self.bg_color), (112, 630))

        # present the place in the lead board
        print(name)
        if name is None:
            msg = "You are not on the lead board"
            prefix = 0
            color = (255, 37, 37)  # red
        else:
            row = 2
            for i in range(2, sheet.max_row + 1):
                if sheet.cell(row=i, column=2).value == name:
                    row = i - 1
            msg = "You are in the {0} place!".format(row)
            if row < 6:  # in the top 5
                color = (50, 190, 60)  # green
            else:
                color = (225, 211, 12)  # yellow
            prefix = 52

        self.screen.blit(self.font.render(msg, False, color, self.bg_color), (112 + prefix, 90))

        pygame.display.update()

        # wait for the user to finish looking at the lead board
        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:  # leave
                    pygame.quit()
                    sys.exit()
                elif event.type == pygame.MOUSEBUTTONDOWN:  # quit lead board
                    return

    def __str__(self):
        return "This is base class for the games"


class FourInARow(Game):
    def __init__(self, screen):
        super().__init__("FourInARow", screen)
        self.cols = 7
        self.rows = 6
        self.game_board = [[0 * x * y for x in range(self.cols)] for y in range(self.rows)]

    def playOneRound(self):
        """
        play one round of the game
        :return: if there was a win, how many turns, if there was a tie
        """
        turn = 0
        user_turn = False
        tie = False
        while not tie:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:  # if user pressed exit, leave game
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEMOTION and user_turn:  # move with the user
                    pygame.draw.rect(self.screen, (0, 0, 0), (0, 0, self.screen_size, 100))
                    pos_x = event.pos[0]
                    pygame.draw.circle(self.screen, (255, 255, 0), (pos_x, int(100 / 2)), 45)
                pygame.display.update()

                if event.type == pygame.MOUSEBUTTONDOWN and user_turn:  # if pressed place the token if it's user's turn
                    pygame.draw.rect(self.screen, (0, 0, 0), (0, 0, self.screen_size, 100))
                    pos_x = event.pos[0]
                    col = int(math.floor(pos_x / 100))
                    msg = self.findPlaceToDrop(col, 2)  # place the token
                    self.drawBoard()  # print the board

                    if "Full" not in msg:  # if succeed to place
                        turn += 1
                        win_msg = self.checkWin()
                        tie = self.checkTie()
                        if "won" in win_msg:  # check for a win
                            sleep(0.75)
                            if "You" in win_msg:
                                self.points += 1
                            return True if "You" in win_msg else False, turn, False
                        elif tie:
                            sleep(0.75)
                            return False, turn, True
                        user_turn = True if not user_turn else False  # switch computer <-> user

                if not user_turn:  # computer's turn
                    sleep_time = random.uniform(0.5, 2)
                    sleep(sleep_time)
                    col = random.randrange(0, 7)
                    msg = self.findPlaceToDrop(col, 1)  # place the token
                    self.drawBoard()  # print the board

                    if "Full" not in msg:  # if succeed to place
                        win_msg = self.checkWin()
                        tie = self.checkTie()
                        if "won" in win_msg:  # check for a win
                            sleep(0.75)
                            return True if "You" in win_msg else False, turn, False
                        elif tie:
                            sleep(0.75)
                            return False, turn, True
                        user_turn = True if not user_turn else False  # switch computer <-> user

        return False, turn, False

    def checkWin(self):
        """
        Check if there is a win and if so, who won - check if there are 4 same symbols in row, column or diagonal
        :return msg - whether there is a win or not:
        """
        # check for win for 4 in row
        for i in range(self.rows):
            curr_player = self.game_board[i][0]
            count_tokens = 1
            for j in range(1, self.cols):
                if curr_player == self.game_board[i][j] and curr_player != 0:  # found token that creates a row
                    count_tokens += 1
                    if count_tokens == 4:  # found 4 tokens in a row
                        msg = "Computer won!\n" if curr_player == 1 else "You won!\n"
                        return msg
                else:  # found different token that breaks the row
                    count_tokens = 1
                    curr_player = self.game_board[i][j]

        # check for win for 4 in column
        for i in range(self.cols):
            curr_player = self.game_board[0][i]
            count_tokens = 1
            for j in range(1, self.rows):
                if curr_player == self.game_board[j][i] and curr_player != 0:  # found token that creates a column
                    count_tokens += 1
                    if count_tokens == 4:  # found 4 tokens in a column
                        msg = "Computer won!\n" if curr_player == 1 else "You won!\n"
                        return msg
                else:  # found different token that breaks the column
                    count_tokens = 1
                    curr_player = self.game_board[j][i]

        # check for win for 4 in diagonal
        for i in range(self.rows):
            for j in range(self.cols):
                try:
                    if self.game_board[i][j] == self.game_board[i + 1][j + 1] == self.game_board[i + 2][j + 2] \
                            == self.game_board[i + 3][j + 3]:  # left->right
                        if self.game_board[i][j] != 0:
                            msg = "Computer won!\n" if self.game_board[i][j] == 1 else "You won!\n"
                            return msg
                except IndexError:
                    break
        for i in range(self.rows):
            for j in range(self.cols):
                try:
                    if self.game_board[i][j] == self.game_board[i - 1][j + 1] == self.game_board[i - 2][j + 2] \
                            == self.game_board[i - 3][j + 3]:  # right->left
                        if self.game_board[i][j] != 0:
                            msg = "Computer won!\n" if self.game_board[i][j] == 1 else "You won!\n"
                            return msg
                except IndexError:
                    break

        return "No win"

    def checkTie(self):
        """
        check if there is a tie - the whole board is full without a winner
        :return True/False - whether there is a tie or not:
        """
        for i in range(self.rows):
            for j in range(self.cols):
                if self.game_board[i][j] == 0:
                    return False
        return True

    def findPlaceToDrop(self, column, player):
        """
        Find the place to put the token to
        :param column: the column to drop in
        :type column: int
        :param player: 1 or 2 - the player number
        :type player: int
        :return msg with the place in the board or msg that says the column is full:
        """
        msg = ""
        for i in range(self.rows):
            if self.game_board[i][column] == 0:
                self.game_board[i][column] = player
                msg = str(i + 1) + ',' + str(column + 1)
                break  # quit the loop
        if msg == "":
            msg = "Full column"
        return msg

    def drawBoard(self):
        """
        draw the board
        :return: None
        """
        self.screen.fill((0, 0, 0))
        pygame.display.update()

        for c in range(self.cols):
            for r in range(self.rows):
                pygame.draw.rect(self.screen, (0, 0, 255), (c * 100, r * 100 + 100, 100, 100))
                pygame.draw.circle(self.screen, (0, 0, 0), (
                    int(c * 100 + 100 / 2), int(r * 100 + 100 + 100 / 2)), 45)

        for c in range(self.cols):
            for r in range(self.rows):
                if self.game_board[r][c] == 1:
                    pygame.draw.circle(self.screen, (255, 0, 0), (
                        int(c * 100 + 100 / 2), 700 - int(r * 100 + 100 / 2)), 45)
                elif self.game_board[r][c] == 2:
                    pygame.draw.circle(self.screen, (255, 255, 0), (
                        int(c * 100 + 100 / 2), 700 - int(r * 100 + 100 / 2)), 45)
        pygame.display.update()

        print()
        for i in range(self.rows - 1, -1, -1):
            print(self.game_board[i])
        print()
        return

    def resetAll(self):
        """
        reset the game parameters
        :return: None
        """
        self.game_board = [[0 * x * y for x in range(self.cols)] for y in range(self.rows)]
        return

    def __str__(self):
        return "This is class for " + self.name


class TicTacToe(Game):
    def __init__(self, screen):
        super().__init__("TicTacToe", screen)
        self.cols = 3
        self.rows = 3
        self.game_board = [[0 * x * y for x in range(self.cols)] for y in range(self.rows)]
        self.x_pic = pygame.image.load(r'photos\TicTacToe\x.png')
        self.o_pic = pygame.image.load(r'photos\TicTacToe\o.png')

    def playOneRound(self):
        """
        play one round of the game
        :return: if there was a win, how many turns, if there was a tie
        """
        turn = 0
        user_turn = False
        win = False
        tie = False
        while not win or not tie:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:  # if user pressed exit, leave game
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEBUTTONDOWN and user_turn:  # if pressed place the token if it's user's turn
                    (pos_x, pos_y) = pygame.mouse.get_pos()
                    valid = self.isValidPlace(pos_x, pos_y, 2)
                    print(pos_x, pos_y)

                    if valid:
                        self.drawBoard()  # print the board

                        # check for win/tie
                        msg_win = self.checkWin()
                        tie = self.checkTie()
                        if "won" in msg_win:
                            sleep(0.75)
                            if "You" in msg_win:
                                self.points += 1
                            return True if "You" in msg_win else False, turn, False
                        elif tie:
                            sleep(0.75)
                            return False, turn, True
                        user_turn = True if not user_turn else False  # switch computer <-> user

                if not user_turn:  # computer's turn
                    turn += 1
                    self.computerChoose()
                    self.drawBoard()  # print the board

                    # check for win/tie
                    msg_win = self.checkWin()
                    tie = self.checkTie()
                    if "won" in msg_win:
                        sleep(0.75)
                        return True if "You" in msg_win else False, turn, False
                    elif tie:
                        sleep(0.75)
                        return False, turn, True
                    user_turn = True if not user_turn else False  # switch computer <-> user

        return True, turn

    def checkWin(self):
        """
        check if there is a win to 'X' or 'O'
        :return: msg who won, string
        """
        players = [1, 2]
        for player in players:
            for i in range(0, 3):  # rows
                if self.game_board[0][i] == self.game_board[1][i] == self.game_board[2][i] == player:
                    return "You won!" if player == 2 else "Computer won!"
            for i in range(0, 3):  # columns
                if self.game_board[i][0] == self.game_board[i][1] == self.game_board[i][2] == player:
                    return "You won!" if player == 2 else "Computer won!"
            # diagonal
            if self.game_board[0][0] == self.game_board[1][1] == self.game_board[2][2] == player \
                    or self.game_board[0][2] == self.game_board[1][1] == self.game_board[2][0] == player:
                return "You won!" if player == 2 else "Computer won!"
        return "No win."

    def checkTie(self):
        """
        check if there is a tie
        :return: True/False - boolean
        """
        for row in self.game_board:
            for place in row:
                if place == 0:
                    return False
        return True

    def isValidPlace(self, pos_x, pos_y, token):
        """
        check if the position that chosen is empty and if so place a token there
        :param pos_x: x part of the position
        :type pos_x: int
        :param pos_y: y part of the position
        :type pos_y: int
        :param token: the token - X or O
        :type token: int
        :return: if empty or not - bool
        """
        # get the row and the column
        if pos_y < self.screen_size / 3:
            row = 0
        elif pos_y < self.screen_size * 2 / 3:
            row = 1
        else:
            row = 2
        if pos_x < self.screen_size / 3:
            col = 0
        elif pos_x < self.screen_size * 2 / 3:
            col = 1
        else:
            col = 2

        # check if empty
        print(self.game_board[row][col])
        if self.game_board[row][col] == 0:
            self.game_board[row][col] = token
            return True
        return False

    def computerChoose(self):
        """
        choose the place for the computer X
        :return:
        """
        sleep_time = random.uniform(0.5, 2)
        sleep(sleep_time)

        start_row = random.randrange(0, 3)
        start_col = random.randrange(0, 3)
        for i in range(start_row, self.rows):
            for j in range(start_col, self.cols):
                if self.game_board[i][j] == 0:  # has to happen, if not - tie should already occur
                    self.game_board[i][j] = 1
                    return
            for j in range(start_col):
                if self.game_board[i][j] == 0:  # has to happen, if not - tie should already occur
                    self.game_board[i][j] = 1
                    return

        for i in range(start_row):
            for j in range(start_col, self.cols):
                if self.game_board[i][j] == 0:  # has to happen, if not - tie should already occur
                    self.game_board[i][j] = 1
                    return
            for j in range(start_col):
                if self.game_board[i][j] == 0:  # has to happen, if not - tie should already occur
                    self.game_board[i][j] = 1
                    return

    def drawBoard(self):
        """
        draw the board
        :return: None
        """
        # draw the clean board
        self.screen.fill((255, 255, 255))

        # drawing vertical lines
        pygame.draw.line(self.screen, (0, 0, 0), (self.screen_size / 3, 0), (self.screen_size / 3, self.screen_size), 7)
        pygame.draw.line(self.screen, (0, 0, 0), (self.screen_size / 3 * 2, 0),
                         (self.screen_size / 3 * 2, self.screen_size), 7)

        # drawing horizontal lines
        pygame.draw.line(self.screen, (0, 0, 0), (0, self.screen_size / 3), (self.screen_size, self.screen_size / 3), 7)
        pygame.draw.line(self.screen, (0, 0, 0), (0, self.screen_size / 3 * 2),
                         (self.screen_size, self.screen_size / 3 * 2), 7)

        pygame.display.update()

        # draw the x and o
        for i in range(self.rows):
            for j in range(self.cols):
                if self.game_board[i][j] == 1:
                    pos = (50 + self.screen_size * j / 3, 50 + self.screen_size * i / 3)
                    self.screen.blit(self.x_pic, pos)
                elif self.game_board[i][j] == 2:
                    pos = (50 + self.screen_size * j / 3, 50 + self.screen_size * i / 3)
                    self.screen.blit(self.o_pic, pos)

        pygame.display.update()

        for row in self.game_board:
            print(row)
        return

    def resetAll(self):
        """
        reset the game board
        :return: None
        """
        self.game_board = [[0 * x * y for x in range(self.cols)] for y in range(self.rows)]

    def __str__(self):
        return "This is class for " + self.name


class Hangman(Game):
    def __init__(self, screen):
        super().__init__("Hangman", screen)
        self.list_words = open(r'files\Hangman\words.txt', 'r').read().split('\n')
        self.secret_word = ""
        self.user_guess = set()
        self.correct_guesses = 0
        self.wrong_guesses = 0
        self.wood = pygame.image.load(r'photos\Hangman\wood.png')
        self.head = pygame.image.load(r'photos\Hangman\head.png')
        self.body = pygame.image.load(r'photos\Hangman\body.png')
        self.hand_left = pygame.image.load(r'photos\Hangman\hand-l.png')
        self.hand_right = pygame.image.load(r'photos\Hangman\hand-r.png')
        self.leg_left = pygame.image.load(r'photos\Hangman\leg-l.png')
        self.leg_right = pygame.image.load(r'photos\Hangman\leg-r.png')

    def playOneRound(self):
        """
        play one round of the game
        :return: if there was a win, how many guesses, if there was a tie
        """
        # initialize the game
        self.secret_word = random.choice(self.list_words)
        self.list_words.remove(self.secret_word)
        self.drawWord()
        self.drawWrong()
        print(self.secret_word)

        # start the game
        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:  # leave
                    pygame.quit()
                    sys.exit()

                elif event.type == pygame.KEYDOWN:  # pressed a key
                    feedback_rect = pygame.Rect(320, 350, 380, 120)
                    pygame.draw.rect(self.screen, self.bg_color, feedback_rect)
                    pygame.display.flip()

                    # generate feedback msg
                    if self.isValidLetter(event.unicode):
                        if event.unicode in self.user_guess:  # already guessed
                            feedback_msg = self.font.render("Already tried that", False, (0, 0, 0), self.bg_color)

                        elif event.unicode in self.secret_word:  # correct guess
                            feedback_msg = self.font.render("Nice!", False, (0, 0, 0), self.bg_color)
                            self.correct_guesses += 1

                        else:  # false guess
                            feedback_msg = self.font.render("Try again!", False, (0, 0, 0), self.bg_color)
                            self.wrong_guesses += 1

                        self.user_guess.add(event.unicode.lower())
                    else:
                        feedback_msg = self.font.render("Only letters", False, (0, 0, 0), self.bg_color)
                    self.screen.blit(feedback_msg, (330, 350))

                # update board
                self.drawWord()
                self.drawWrong()
                self.drawHangman()
                pygame.display.update()

                # check if won
                win, who = self.checkWin()
                if win and who == "user":
                    feedback_rect = pygame.Rect(320, 350, 380, 120)
                    pygame.draw.rect(self.screen, self.bg_color, feedback_rect)
                    self.screen.blit(self.font.render("The word was " + self.secret_word,
                                                      False, (0, 0, 0), self.bg_color), (330, 350))
                    pygame.display.flip()
                    sleep(0.75)
                    self.points += 6 - self.wrong_guesses
                    return True, len(self.user_guess), False
                elif win:
                    feedback_rect = pygame.Rect(320, 350, 380, 120)
                    pygame.draw.rect(self.screen, self.bg_color, feedback_rect)
                    self.screen.blit(self.font.render("The word was " + self.secret_word,
                                                      False, (0, 0, 0), self.bg_color), (330, 350))
                    pygame.display.flip()
                    sleep(0.75)
                    return False, len(self.user_guess), False

    def checkWin(self):
        """
        check if the user won or the computer (which mean the user lost)
        :return: True if there was a win (to the user/computer) or False if no one won, and string with whom won
        """
        secret_letters = set()  # in set to avoid double letters
        for letter in self.secret_word:
            secret_letters.add(letter)

        if self.wrong_guesses == 6:
            return True, "computer"
        elif self.correct_guesses == len(secret_letters):
            return True, "user"
        return False, "no one"

    @staticmethod
    def isValidLetter(letter):
        """
        check if the letter the user entered is valid (a letter in english)
        :param letter: the letter the user entered
        :type letter: str
        :return: True or False
        """
        if not letter.isalpha():
            return False
        return True

    def drawBoard(self):
        """
        draw the board of the game in each turn
        :return: None
        """
        # draw the screen
        self.screen.fill(self.bg_color)

        # headings
        self.screen.blit(pygame.font.SysFont("arial bold", 80).render("Try to guess!",
                                                                      False, (0, 0, 0), self.bg_color), (315, 100))
        self.screen.blit(self.font.render("Type the letters", False, (0, 0, 0), self.bg_color), (320, 165))

        # the hangman
        self.screen.blit(self.wood, (5, 70))
        pygame.display.update()
        return

    def drawWord(self):
        """
        draw the guesses and the empty places of the word needed to be guessed
        :return: None
        """
        # the guess word
        guess_bars_txt = ""
        for letter in self.secret_word:
            if letter.lower() in self.user_guess:
                guess_bars_txt += letter
            else:
                guess_bars_txt += "_"
            guess_bars_txt += "  "

        guess_bars = pygame.font.SysFont("arial bold", 80).render(guess_bars_txt, False, (0, 0, 0), self.bg_color)
        self.screen.blit(guess_bars, (330, 280))
        pygame.display.update()

    def drawWrong(self):
        """
        present the wrong letter that have been guessed
        :return: None
        """
        # wrong guesses
        self.screen.blit(self.font.render("Wrong guesses:", False, (0, 0, 0), self.bg_color), (30, 515))
        wrong_txt = ""
        for letter in self.user_guess:
            if letter not in self.secret_word and letter.upper() not in self.secret_word:
                wrong_txt += letter
                wrong_txt += "  "

        self.screen.blit(self.font.render(wrong_txt, False, (0, 0, 0), self.bg_color), (30, 570))
        pygame.display.update()
        return

    def drawHangman(self):
        """
        draw the hangman according to the number of mistakes
        :return: None
        """
        # draw the man if the user was wrong
        if self.wrong_guesses > 0:
            self.screen.blit(self.head, (173, 198))
        if self.wrong_guesses > 1:
            self.screen.blit(self.body, (185, 261))
        if self.wrong_guesses > 2:
            self.screen.blit(self.hand_left, (164, 267))
        if self.wrong_guesses > 3:
            self.screen.blit(self.hand_right, (211, 267))
        if self.wrong_guesses > 4:
            self.screen.blit(self.leg_left, (175, 346))
        if self.wrong_guesses > 5:
            self.screen.blit(self.leg_right, (207, 346))
        return

    def resetAll(self):
        """
        reset all the data
        :return: None
        """
        self.user_guess = set()
        self.wrong_guesses = 0
        self.correct_guesses = 0
        self.secret_word = ""
        return

    def statistics_round(self, won, turns, round_time, choosing, curr_round, tie):
        """
        calculate and present the statistics of each round of the game
        :param won: whether the user won the round
        :type won: bool
        :param turns: number of turns this round took
        :type turns: int
        :param round_time: the time the round took
        :type round_time: float
        :param choosing: the average time it took the user to choose an action
        :type choosing: float
        :param curr_round: number of the current round
        :type curr_round: int
        :param tie: whether the round ended with a tie
        :type tie: bool
        :return: the msg of the statistics, string
        """
        self.stat['time'].append(round_time)  # save time
        self.stat['turns'].append(turns)  # save turns
        self.stat['choosing'].append(choosing)
        if won:
            self.stat['wins'] += 1  # remember wins of user
        if tie:
            self.stat['ties'] += 1  # remember ties

        avg_wins = self.stat['wins'] / curr_round  # calc avg wins per round
        avg_turns = sum(self.stat['turns']) / curr_round  # calc avg turns per round

        # make msg after one round
        msg = "Here are the round statistics:\n"
        msg += "You guessed right: " + str(self.stat['wins']) + " times\n"
        msg += "You guessed wrong: " + str(curr_round - self.stat['ties'] - self.stat['wins']) + " times\n"
        msg += "Amount of guesses for this round: " + str(turns) + "\n"
        msg += "This round took you " + str(format(self.stat['time'][curr_round - 1], '.3f')) + " seconds\n"
        # msg += "You chose where to place the token in " + str(format(self.stat['choosing'][curr_round - 1], '.3f'))
        # msg += " seconds in average\n"
        msg += "Average amount of guesses per round: " + str(format(avg_turns, '.3f')) + "\n"
        msg += "Average amount of wins per round: " + str(format(avg_wins, '.3f')) + "\n\n"

        # present statistics
        self.screen.fill(self.bg_color)
        pygame.display.update()
        sleep(0.5)
        stat_msg = pygame.font.SysFont("arial bold", 60).render("Round ended!", False, (0, 0, 0), self.bg_color)
        self.screen.blit(stat_msg, (210, 20))

        # win msg
        if won:
            msg_win = "You are right! Congratulations!"
            msg_win_color = (50, 190, 60)
            prefix = 0

        else:
            msg_win = "You guessed wrong! Try again."
            msg_win_color = (255, 37, 37)
            prefix = 11
        stat_msg = pygame.font.SysFont("arial bold", 60).render(msg_win, False, msg_win_color, self.bg_color)
        self.screen.blit(stat_msg, (32 + prefix, 65))

        # points msg
        point_msg = pygame.font.SysFont("arial bold", 50).render("You've got {0} points!".format(self.points),
                                                                 False, (0, 0, 0), self.bg_color)
        self.screen.blit(point_msg, (185, 115))

        # the stats
        for i in range(len(msg.split('\n')) - 1):
            stat_msg = pygame.font.SysFont("arial", 38).render(msg.split('\n')[i], False, (0, 0, 0), self.bg_color)
            k = 10 if i == 0 else 50
            self.screen.blit(stat_msg, (k, i * 50 + 172))

        self.screen.blit(self.font.render("Press the mouse to continue", False, (0, 0, 0), self.bg_color), (112, 550))
        pygame.display.flip()  # display the msg

        while True:  # present until pressed mouse
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEBUTTONDOWN:
                    break
            else:
                continue
            break
        print(msg)
        return msg

    def statistics_game(self, rounds):
        """
        calculate and present the statistics of the whole game
        :param rounds: number of rounds the game took
        :type rounds: int
        :return: the msg of the statistics, string
        """
        avg_wins = self.stat['wins'] / rounds  # calc avg wins per round
        avg_turns = sum(self.stat['turns']) / rounds  # calc avg turns per round
        avg_time = sum(self.stat['time']) / rounds  # calc avg time per round
        msg = "Here are the game statistics:\n"
        msg += "You guessed right: " + str(self.stat['wins']) + " times\n"
        msg += "You guessed wrong: " + str(rounds - self.stat['wins']) + " times\n"
        msg += "Average amount of time per round: " + str(format(avg_time, '.3f')) + "s\n"
        # msg += "You choose where to place the token in " + str(format(self.stat['choosing'][rounds - 1], '.3f'))
        # msg += " seconds in average\n"
        msg += "Average amount of guesses per round: " + str(format(avg_turns, '.3f')) + "\n"
        msg += "Average amount of wins per round: " + str(format(avg_wins, '.3f')) + "\n\n"

        # present statistics
        self.screen.fill(self.bg_color)
        pygame.display.update()
        sleep(0.5)

        # the stats
        stat_msg = pygame.font.SysFont("arial bold", 60).render("Game ended!", False, (0, 0, 0), self.bg_color)
        self.screen.blit(stat_msg, (210, 20))

        # points msg
        point_msg = pygame.font.SysFont("arial bold", 50).render("You've got {0} points!".format(self.points),
                                                                 False, (0, 0, 0), self.bg_color)
        self.screen.blit(point_msg, (185, 70))

        # stat msg
        for i in range(len(msg.split('\n')) - 1):
            stat_msg = pygame.font.SysFont("arial", 38).render(msg.split('\n')[i], False, (0, 0, 0), self.bg_color)
            k = 10 if i == 0 else 50
            self.screen.blit(stat_msg, (k, i * 50 + 127))

        self.screen.blit(self.font.render("Press the mouse to continue", False, (0, 0, 0), self.bg_color), (112, 550))
        pygame.display.flip()  # display the msg

        while True:  # present until pressed mouse
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

                if event.type == pygame.MOUSEBUTTONDOWN:
                    break
            else:
                continue
            break
        print(msg)
        return msg

    def __str__(self):
        return "This is class for " + self.name
